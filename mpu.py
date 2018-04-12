#!/usr/bin/python
# coding: utf-8 -*-

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl import __version__
from pandas import ExcelWriter


wb = openpyxl.Workbook()
wb.save('tables.xlsx')
book = load_workbook('tables.xlsx')
writer = ExcelWriter('tables.xlsx', engine='openpyxl') 	
writer.book = book
writer.sheets = dict((ws.title, ws) for ws in book.worksheets) 


class XmlFile():

#Methods for respectve table

    def table_0(self,L0,L1):
         df = pd.DataFrame({"Emp_ID":L0,"Employee":L1})
         #index=False flag remove default 1st column with indices
         df.to_excel(writer,sheet_name='SampleTable',index=False)
         writer.save()

    def table_1(self,L0,L1,L2):
         df = pd.DataFrame({"aggregateMID":L0,"MID":L1,"sequenceNum":L2})
         df.to_excel(writer,sheet_name='aggregateTable',index=False)
         writer.save()

    def table_2(self,L0,L1,L2,L3,L4,L5,L6,L7,L8,L9,L10):
         df = pd.DataFrame({"MID":L0,"LID":L1,"title":L2,"shortTitle":L3,"artist":L4,"genre":L5,"description":L6,"shortDescription":L7,"criticScore":L8,"year":L9,"copyright":L10})
         df.to_excel(writer,sheet_name='audioDescriptionTable',index=False)
         writer.save()

    def table_3(self,L0,L1,L2,L3,L4,L5,L6):
         df = pd.DataFrame({"mediaConfigId":L0,"startDate":L1,"endDate":L2,"class":L3,"CID":L4,"sequenceNum":L5,"parentCID":L6})
         df.to_excel(writer,sheet_name='categoryConfigTable',index=False)
         writer.save()

    def table_4(self,L0,L1,L2,L3,L4,L5,L6):
         df = pd.DataFrame({"CID":L0,"LID":L1,"title":L2,"description":L3,"shortDescription":L4,"synopsisImgID":L5,"posterImgID":L6})
         df.to_excel(writer,sheet_name='categoryInfoTable',index=False)
         writer.save()

    def table_5(self,L0,L1,L2,L3,L4):
         df = pd.DataFrame({"MID":L0,"LID":L1,"ccSubtitleDef":L2,"PID":L3,"languageOrder":L4})
         df.to_excel(writer,sheet_name='ccSubtitlesTable',index=False)
         writer.save()

    def table_6(self,L0,L1,L2,L3,L4,L5,L6,L7,L8):
         df = pd.DataFrame({"MID":L0,"LID":L1,"title":L2,"shortTitle":L3,"director":L4,"cast":L5,"year":L6,"genre":L7,"description":L8})
         df.to_excel(writer,sheet_name='videoDescriptionTable',index=False)
         writer.save()



